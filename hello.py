#-*- coding: UTF-8 -*- 
from flask import Flask,render_template,request,redirect,session,url_for
from flask import send_file, send_from_directory,app
from openpyxl import Workbook
import sys, os
from flask_bootstrap import Bootstrap
from flask_wtf import FlaskForm
from wtforms import TextAreaField, SubmitField
from wtforms.fields.html5 import EmailField
from threading import Thread
from wtforms.validators import DataRequired,InputRequired,Email
#fuzzywuzzy 注意重名
#from .fuzzycompare import compare2list 包引用重构时候再说
from fuzzywuzzy import process as fuzzyprocess,fuzz 
import logging


# Flask 用这个参数确定应用的位置，进而找到应用中其他文件的位置，例如图像和模板。
app = Flask(__name__)

#using fuzzywuzzy
def compare2list(leftl,rightl,resultdict):
    '''
    2 list, using right one to match left one 
    '''
    for item in range(len(leftl)):
        resultdict[item]=[leftl[item],fuzzyprocess.extractOne(leftl[item], rightl,scorer=fuzz.token_sort_ratio)]


#数据库
from flask_sqlalchemy import SQLAlchemy
basedir = os.path.abspath(os.path.dirname(__file__))
app.config['SQLALCHEMY_DATABASE_URI'] ='sqlite:///' + os.path.join(basedir, 'data.sqlite')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)


class User(db.Model):
    __tablename__ = 'users'
    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(64), unique=True)
    cnt= db.Column(db.Integer)
    def __repr__(self):
        return '<Role %r>' % self.name



sys.path.append(os.getcwd())
#表单key
bootstrap = Bootstrap(app)
app.config['SECRET_KEY'] = 'fuzzyflask' #表单类使用为了增强安全性，密钥不应该直接写入源码，而要保存在环境变量中。这一技术在第 7 章介绍。

#邮件
from flask_mail import Mail

app.config.update(dict(
    DEBUG = True,
    MAIL_SERVER = 'smtp.qq.com',#'smtp.gmail.com',
    MAIL_PORT = '465',#
    MAIL_USE_TLS = False,#587
    MAIL_USE_SSL = True,#465
    MAIL_USERNAME = '1274543351@qq.com',#os.environ.get('MAIL_USERNAME'),
    MAIL_PASSWORD = 'lvjgwyfpbzatgibc',# os.environ.get('MAIL_PASSWORD') export MAIL_PASSWORD=
    MAIL_DEFAULT_SENDER	= '1274543351@qq.com',
))

mail = Mail(app)#connection refused

from flask_mail import Message
app.config['FLASKY_MAIL_SUBJECT_PREFIX'] = '[matchfuzzydata]'
app.config['FLASKY_MAIL_SENDER'] = 'matchfuzzydata <1274543351@qq.com>'

#send_email() 函数的参数分别为收件人地址、主题、渲染邮件正文的模板和关键字参数列表。
# 调用者传入的关键字参数将传给 render_template() 函数，作为模板变量提供给
# 模板使用，用于生成电子邮件正文。 todo我们可以把执行 send_async_email() 函数的操作发给 Celery 任务队列。
def send_email(to, subject, template, **kwargs):
    msg = Message(app.config['FLASKY_MAIL_SUBJECT_PREFIX'] + subject,sender=app.config['FLASKY_MAIL_SENDER'], recipients=[to])
    msg.body = render_template(template + '.txt', **kwargs)
    msg.html = render_template(template + '.html', **kwargs)
    with app.open_resource("static/"+to+".xlsx") as fp:
        msg.attach("static/"+to+".xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", fp.read())
    thr = Thread(target=send_async_email, args=[app, msg])
    thr.start()
    return thr

def send_async_email(app, msg):
    with app.app_context():
        mail.send(msg)


#@app.route('/') 主页
# 定义路由的最简便方式，是使用应用实例提供的app.route 装饰器
@app.route('/',methods=['GET','POST'] )
def index():
    form = NameForm()#表单模型初始化
    if form.validate_on_submit(): #刷新不重复提交post,这里没啥用……
        session['leftl'] = form.leftl.data
        session['rightl'] = form.leftl.data
        return redirect(url_for('index'))
    #print("show index")
    return render_template('index.html',form=form,leftl=session.get('leftl'), \
        rightl=session.get('rightl'),email=session.get('email'),submit=session.get('submit'))


#about页面
@app.route('/about',methods=['GET','POST'] )
def about():
    return render_template('about.html')


#主程序，邮件版本
@app.route('/process',methods=['GET','POST'] )
def process():
    print("start process")
    form = NameForm()
    #string to list flask处理request的方法不一样
    leftl=request.form.get('leftl').split("\r\n")
    rightl=request.form.get('rightl').split("\r\n")
    email=form.email.data
    resultdic={}
    compare2list(leftl,rightl,resultdic)
    #生成excel
    wb = Workbook()
    ws = wb.active
    #表头
    ws.cell(column=1 , row=1, value="顺序")
    ws.cell(column=2 , row=1, value="非标准名称")
    ws.cell(column=3 , row=1, value="标准名称")
    ws.cell(column=4 , row=1, value="相似度")
    next_row=2
    #依次生成
    for (key,value) in resultdic.items():
        ws.cell(column=1 , row=next_row, value=key)
        ws.cell(column=2 , row=next_row, value=value[0])
        ws.cell(column=3 , row=next_row, value=value[1][0])
        ws.cell(column=4 , row=next_row, value=value[1][1])
        next_row += 1

    #数据库处理邮件地址，判断使用次数 3次
    #print(email)
    user=User.query.filter_by(email=email).first()
    if user is None:#新客户
        user=User(email=email)
        db.session.add(user)
        user.cnt=1
        db.session.add(user)
        db.session.commit()
        #sendfile(address)
        # downloadfile(address)
    elif user.cnt<3:
        user.cnt+=1
        db.session.add(user)
        db.session.commit()
        #sendfile(address)
        # downloadfile(address)
    else:#超过3次,付费后send file
        user.cnt+=1
        db.session.add(user)
        db.session.commit()
        #print("first wechat payment part")
        #sendfile(address)
        # downloadfile(address)
        #todo 付款环节
    #返回处理后的结果 ，更新form显示，处理结果已经生成，等待下载
    #按客户邮箱命名文件
    filename=email+str(user.cnt)+".xlsx"
    wb.save("donefiles/"+filename)
    print(url_for('process',_external=True)+'/'+filename)

    #return render_template('results.html',form=form,leftl=leftl,rightl=rightl,email=email)

    #本地测试下载——最后直接用nginx了
    # return send_file("donefiles/"+filename,
    #                  mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    #                  attachment_filename=filename,
    #                  as_attachment=True)

    #nginx在线返回静态附件url
    #url_for('user', name='john', _external=True) 的返回结果是 http://localhost:5000/user/john
    #return redirect("youtube.com")

    return redirect(url_for('process',_external=True)+'/'+filename)



#快速生成bootstrap风格表单
class NameForm(FlaskForm):
    leftl = TextAreaField('待匹配的非标准名称列', validators=[DataRequired()])
    rightl = TextAreaField('用来匹配的标准名称列', validators=[DataRequired()])
    email = EmailField("你的邮箱~",  validators=[InputRequired("Please enter your email address"), Email("Please enter your email address.")])
    submit = SubmitField('发送匹配结果:)')
    downloadfile=SubmitField('下载结果')


@app.route('/user/<name>')#动态路由获得参数例子
def user(name):
    return render_template('user.html', name=name)

@app.errorhandler(404)
def page_not_found(e):
    return render_template('404.html'),404

@app.errorhandler(500)
def internal_server_error(e):
    return render_template('500.html'), 500



if __name__ == '__main__':
    app.run()